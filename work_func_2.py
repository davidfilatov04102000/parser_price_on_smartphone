from openpyxl import load_workbook
from smartphones import smartphone_list
import os


def main_func_2(arg_list):
    list_new_price_model_for_rewrite = []
    new_list_model = []
    new_list_price = []
    list_changes_price = []
    list_old_price_for_versus = []

    list_all_model = arg_list.copy()
    set_list = set(list_all_model)

    list_no_repetitions = list(set_list).copy()

    rp = open("place_models.txt", "w")
    for ret in list_all_model:
        rp.write(ret + "\n")
    rp.close()

    def rename_remove_file():
        cirilic_list = ['а', 'б', 'в', 'г', 'д', 'е', 'ё', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с',
                        'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ъ', 'ы', 'ь', 'э', 'ю', 'я']

        for x in os.listdir(path="."):
            if x[1] in cirilic_list:
                os.remove("data.xlsx")
                os.rename(x, "data.xlsx")

    def parsing_xlsx():
        wb = load_workbook(filename = r'data.xlsx', data_only=True)

        ws = wb.worksheets[3]

        for x in range(3, 82+1):
            model = str(ws[f"A{x}"].value)
            price_12 = str(ws[f"E{x}"].value)
            price_24 = str(ws[f"I{x}"].value)
            if model in list_no_repetitions:
                res_for_list_for_rewrite = model + ", " + price_12[:7] + ", " + price_24[:7]
                list_new_price_model_for_rewrite.append(res_for_list_for_rewrite)
                res_list_model = model
                new_list_model.append(res_list_model)
                res_list_price = price_12[:7] + ", " + price_24[:7]
                new_list_price.append(res_list_price)

    def rewrite_file():
        with open("Models.txt", "w") as f:
            f.write("")

        with open("price_12_24.txt", "w") as a:
            a.write("")

        for y in list_new_price_model_for_rewrite:
            res_list = y.split(", ")
            with open("Models.txt", "a") as p:
                p.writelines(res_list[0] + "\n")
            with open("price_12_24.txt", "a") as t:
                res_price = res_list[1] + ", " + res_list[2] + "\n"
                t.writelines(res_price)

    rename_remove_file()
    parsing_xlsx()
    rewrite_file()

    list_for_output_bots = []

    for d in list_all_model:
        shet = 0
        for ddr in new_list_model:
            if d == ddr:
                res = ddr + ", " + new_list_price[shet]
                list_for_output_bots.append(res)
            shet += 1

    return list_for_output_bots
