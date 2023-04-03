from openpyxl import load_workbook
from smartphones import smartphone_list
import os


def main_func(arg_list):
    new_list = []
    list_changes_price = []
    list_old_price_for_versus = []

    list_all_model = arg_list.copy()
    set_list = set(list_all_model)

    list_no_repetitions = list(set_list).copy()

    rp = open("place_models.txt", "w")
    for ret in list_all_model:
        rp.write(ret + "\n")
    rp.close()


    def rename_remove_file():
        cirilic_list = ['а', 'б', 'в', 'г', 'д', 'е', 'ё', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с',
                        'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ъ', 'ы', 'ь', 'э', 'ю', 'я']

        for x in os.listdir(path="."):
            if x[1] in cirilic_list:
                os.remove("data.xlsx")
                os.rename(x, "data.xlsx")


    def parsing_xlsx():
        wb = load_workbook(filename = r'data.xlsx', data_only=True)

        ws = wb.worksheets[3]

        for x in range (3, 82+1):
            model = str(ws[f"A{x}"].value)
            price_12 = str(ws[f"E{x}"].value)
            price_24 = str(ws[f"I{x}"].value)
            if model in list_no_repetitions:
                res = model + ", " + price_12[:7] + ", " + price_24[:7]
                new_list.append(res)


    def price_analize():
        mod_list_fr_out = []
        price_list_fr_out = []

        fp = open('Models.txt')
        for line in fp:
            mod_list_fr_out.append(line[:-1])
        fp.close()

        at = open("price_12_24.txt")
        for line in at:
            res_price = line[:-1].split(", ")
            price_list_fr_out.append(res_price)
        at.close()

        price_dicts = dict(zip(mod_list_fr_out, price_list_fr_out))

        for w in new_list:
            timing_list = w.split(", ")
            for key in price_dicts.keys():
                if key in timing_list[0]:
                    if timing_list[1] == price_dicts[key][0]:
                        continue
                    elif timing_list[2] == price_dicts[key][1]:
                        continue
                    else:
                        quan_model = list_all_model.count(timing_list[0])
                        timing_list.append(quan_model)
                        list_changes_price.append(timing_list)
                        res_old_price = key + ", " + price_dicts[key][0] + ", " + price_dicts[key][1]
                        list_res_old_price = res_old_price.split(", ")
                        list_old_price_for_versus.append(list_res_old_price)


    def rewrite_file():
        with open("Models.txt", "w") as f:
            f.write("")

        with open("price_12_24.txt", "w") as a:
            a.write("")

        for y in new_list:
            res_list = y.split(", ")
            with open("Models.txt", "a") as p:
                p.writelines(res_list[0] + "\n")
            with open("price_12_24.txt", "a") as t:
                res_price = res_list[1] + ", " + res_list[2] + "\n"
                t.writelines(res_price)
        # new_list.clear()
        # list_changes_price.clear()
        # list_old_price_for_versus.clear()


    read_file_1 = []
    read_file_2 = []

    fp = open("Models.txt", "r")
    for l in fp:
        read_file_1.append(l)
    fp.close()

    dr = open("price_12_24.txt")
    for h in dr:
        read_file_2.append(h)
    dr.close()

    len_list = len(read_file_1) + len(read_file_2)

    if len_list == 0:
        rename_remove_file()
        parsing_xlsx()
        rewrite_file()
        print(new_list)
        return new_list
    else:
        rename_remove_file()
        parsing_xlsx()
        price_analize()
        rewrite_file()
        print("цена изменилась на следующих моделях", list_changes_price)
        print("старые цены", list_old_price_for_versus)
        print(new_list)
        return list_changes_price



